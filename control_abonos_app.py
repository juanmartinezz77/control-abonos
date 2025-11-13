# Production app — fixes for widget session_state usage and clearing inputs after submit (callbacks)
import re
import streamlit as st
import sqlite3
import pandas as pd
import logging
from io import BytesIO
from datetime import date, datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ------------------ Config / Logging ------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[logging.StreamHandler()],
)

DB_FILENAME_TEMPLATE = "control_abonos_{user}.db"

# ------------------ Helpers DB per user ------------------


def sanitize_username(username: str) -> str:
    if not username:
        return "anonymous"
    return re.sub(r"[^A-Za-z0-9_-]", "_", username)


def get_db_path_for_user(username: str) -> str:
    safe = sanitize_username(username)
    return DB_FILENAME_TEMPLATE.format(user=safe)


def get_connection_for_user(username: str):
    db_path = get_db_path_for_user(username)
    try:
        conn = sqlite3.connect(db_path, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        return conn
    except sqlite3.Error:
        logging.exception("Error conectando a la DB para usuario %s", username)
        st.error("Error al conectar con la base de datos. Revisa los logs.")
        st.stop()


def ensure_column(conn, table: str, column: str, col_type: str):
    c = conn.cursor()
    info = c.execute(f"PRAGMA table_info({table})").fetchall()
    columns = [row[1] for row in info]
    if column not in columns:
        try:
            c.execute(f"ALTER TABLE {table} ADD COLUMN {column} {col_type}")
            conn.commit()
            logging.info("Added column %s to %s", column, table)
        except Exception:
            logging.exception("No se pudo añadir la columna %s a %s", column, table)


def init_db(conn):
    c = conn.cursor()
    c.execute(
        """
        CREATE TABLE IF NOT EXISTS casos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cliente TEXT NOT NULL,
            descripcion TEXT,
            valor_acordado REAL NOT NULL DEFAULT 0,
            etapa TEXT,
            observaciones TEXT,
            creado_en TEXT DEFAULT (DATE('now'))
        )
        """
    )
    c.execute(
        """
        CREATE TABLE IF NOT EXISTS abonos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            monto REAL NOT NULL,
            caso_id INTEGER NOT NULL,
            observaciones TEXT,
            creado_en TEXT DEFAULT (DATE('now')),
            FOREIGN KEY(caso_id) REFERENCES casos(id) ON DELETE CASCADE
        )
        """
    )
    conn.commit()
    ensure_column(conn, "casos", "creado_por", "TEXT")
    ensure_column(conn, "abonos", "creado_por", "TEXT")


# ------------------ CRUD ------------------


def fetch_casos(conn, cliente_filter=None, etapa_filter=None):
    q = "SELECT * FROM casos"
    params, conditions = [], []
    if cliente_filter and cliente_filter != "Todos":
        conditions.append("cliente = ?")
        params.append(cliente_filter)
    if etapa_filter and etapa_filter != "Todos":
        conditions.append("etapa = ?")
        params.append(etapa_filter)
    if conditions:
        q += " WHERE " + " AND ".join(conditions)
    q += " ORDER BY id"
    return pd.read_sql_query(q, conn, params=params)


def fetch_abonos(conn, caso_id=None):
    q = """SELECT abonos.*, casos.cliente, casos.descripcion
           FROM abonos JOIN casos ON abonos.caso_id = casos.id"""
    params = []
    if caso_id:
        q += " WHERE caso_id = ?"
        params.append(caso_id)
    q += " ORDER BY fecha DESC, id DESC"
    df = pd.read_sql_query(q, conn, params=params)
    # normalize fecha to date-only for display (store may contain date string)
    if not df.empty and "fecha" in df.columns:
        try:
            df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce").dt.date
        except Exception:
            logging.exception("No se pudo convertir columna fecha a date")
    return df


def add_caso(conn, cliente, descripcion, valor_acordado, etapa, observaciones, creado_por=None):
    if not cliente or str(cliente).strip() == "":
        raise ValueError("El nombre del cliente es obligatorio.")
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM casos WHERE cliente = ? AND descripcion = ?", (cliente, descripcion))
    if c.fetchone()[0] > 0:
        raise ValueError("Ya existe un caso con ese cliente y descripción.")
    created_date = datetime.utcnow().date().isoformat()  # date-only
    c.execute(
        "INSERT INTO casos (cliente, descripcion, valor_acordado, etapa, observaciones, creado_en, creado_por) VALUES (?,?,?,?,?,?,?)",
        (cliente.strip(), descripcion, float(valor_acordado or 0), etapa, observaciones, created_date, creado_por),
    )
    conn.commit()
    logging.info("Caso agregado: %s - %s (por %s)", cliente, descripcion, creado_por)
    return c.lastrowid


def edit_caso(conn, caso_id, cliente, descripcion, valor_acordado, etapa, observaciones):
    c = conn.cursor()
    c.execute(
        "UPDATE casos SET cliente=?, descripcion=?, valor_acordado=?, etapa=?, observaciones=? WHERE id=?",
        (cliente, descripcion, float(valor_acordado or 0), etapa, observaciones, caso_id),
    )
    conn.commit()
    logging.info("Caso editado id=%s", caso_id)
    return c.rowcount


def delete_caso(conn, caso_id):
    c = conn.cursor()
    c.execute("DELETE FROM abonos WHERE caso_id = ?", (caso_id,))
    c.execute("DELETE FROM casos WHERE id = ?", (caso_id,))
    conn.commit()
    logging.info("Caso eliminado id=%s", caso_id)


def add_abono(conn, fecha, monto, caso_id, observaciones, creado_por=None):
    c = conn.cursor()
    try:
        caso_id_int = int(caso_id)
    except Exception:
        raise ValueError("ID de caso inválido.")
    c.execute("SELECT 1 FROM casos WHERE id = ?", (caso_id_int,))
    if c.fetchone() is None:
        raise ValueError(f"No existe el caso con id {caso_id_int}.")
    try:
        monto_val = float(monto)
    except Exception:
        raise ValueError("Monto inválido.")
    if monto_val <= 0:
        raise ValueError("El monto debe ser mayor que cero.")
    if isinstance(fecha, date):
        fecha_iso = fecha.isoformat()
    else:
        try:
            fecha_iso = pd.to_datetime(fecha).date().isoformat()
        except Exception:
            fecha_iso = datetime.utcnow().date().isoformat()
    created_date = datetime.utcnow().date().isoformat()
    c.execute(
        "INSERT INTO abonos (fecha, monto, caso_id, observaciones, creado_en, creado_por) VALUES (?,?,?,?,?,?)",
        (fecha_iso, monto_val, caso_id_int, observaciones, created_date, creado_por),
    )
    conn.commit()
    logging.info("Abono agregado: caso_id=%s monto=%s fecha=%s por=%s", caso_id_int, monto_val, fecha_iso, creado_por)
    return c.lastrowid


def edit_abono(conn, abono_id, fecha, monto, caso_id, observaciones):
    c = conn.cursor()
    if isinstance(fecha, date):
        fecha_iso = fecha.isoformat()
    else:
        try:
            fecha_iso = pd.to_datetime(fecha).date().isoformat()
        except Exception:
            fecha_iso = datetime.utcnow().date().isoformat()
    c.execute(
        "UPDATE abonos SET fecha=?, monto=?, caso_id=?, observaciones=? WHERE id=?",
        (fecha_iso, float(monto), int(caso_id), observaciones, int(abono_id)),
    )
    conn.commit()
    logging.info("Abono editado id=%s", abono_id)
    return c.rowcount


def delete_abono(conn, abono_id):
    c = conn.cursor()
    c.execute("DELETE FROM abonos WHERE id = ?", (abono_id,))
    conn.commit()
    logging.info("Abono eliminado id=%s", abono_id)


# ------------------ Reports / Exports ------------------


def resumen_por_caso(conn, cliente_filter=None, etapa_filter=None):
    casos = fetch_casos(conn, cliente_filter, etapa_filter)
    if casos.empty:
        return pd.DataFrame(columns=["id", "cliente", "descripcion", "valor_acordado", "total_abonado", "saldo_pendiente", "etapa", "observaciones"])
    abonos = pd.read_sql_query("SELECT caso_id, SUM(monto) as total_abonado FROM abonos GROUP BY caso_id", conn)
    merged = casos.merge(abonos, left_on="id", right_on="caso_id", how="left")
    merged["total_abonado"] = merged["total_abonado"].fillna(0.0)
    merged["saldo_pendiente"] = merged["valor_acordado"] - merged["total_abonado"]
    result = merged[["id", "cliente", "descripcion", "valor_acordado", "total_abonado", "saldo_pendiente", "etapa", "observaciones"]].copy()
    result["valor_acordado"] = result["valor_acordado"].astype(float)
    result["total_abonado"] = result["total_abonado"].astype(float)
    result["saldo_pendiente"] = result["saldo_pendiente"].astype(float)
    return result


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Resumen")
        ws = writer.sheets["Resumen"]
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(border_style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 4, 60)
        for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
            if all((isinstance(c.value, (int, float)) or c.value is None) for c in col):
                for cell in col:
                    cell.number_format = "#,##0.00"
    buffer.seek(0)
    return buffer.read()


# ------------------ UI Helpers ------------------


def money(v):
    try:
        return f"${float(v):,.2f}"
    except Exception:
        return v


# ------------------ Auth ------------------


def check_password(user: str, password: str) -> bool:
    if "credentials" not in st.secrets:
        st.error("Aplicación no configurada: falta la sección [credentials] en los secretos.")
        return False
    creds = st.secrets["credentials"]
    stored = None
    try:
        if hasattr(creds, "__contains__") and user in creds:
            stored = creds[user]
    except Exception:
        stored = None
    if stored is None:
        try:
            stored = getattr(creds, user)
        except Exception:
            stored = None
    if stored is None and hasattr(creds, "get"):
        try:
            stored = creds.get(user)
        except Exception:
            stored = None
    if isinstance(stored, str):
        return password == stored
    if stored is not None:
        try:
            if hasattr(stored, "get"):
                pw = stored.get("password", None)
                if pw is not None:
                    return pw == password
        except Exception:
            pass
        try:
            pw = getattr(stored, "password", None)
            if pw is not None:
                return pw == password
        except Exception:
            pass
    return False


# ------------------ Callbacks for on_click (safe for session_state modifications) ------------------


def submit_new_case(usuario: str):
    try:
        conn = get_connection_for_user(usuario)
        init_db(conn)
        add_caso(
            conn,
            st.session_state.get("new_cliente", ""),
            st.session_state.get("new_descripcion", ""),
            st.session_state.get("new_valor", 0.0),
            st.session_state.get("new_etapa", ""),
            st.session_state.get("new_obs", ""),
            creado_por=usuario,
        )
        # clear state keys (safe inside callback)
        st.session_state["new_cliente"] = ""
        st.session_state["new_valor"] = 0.0
        st.session_state["new_descripcion"] = ""
        st.session_state["new_etapa"] = ""
        st.session_state["new_obs"] = ""
        st.session_state["feedback"] = "Caso agregado correctamente."
        try:
            st.experimental_rerun()
        except Exception:
            pass
    except Exception as e:
        logging.exception("Error agregando caso (callback)")
        st.session_state["feedback"] = f"Error agregando caso: {e}"


def submit_new_abono(usuario: str):
    try:
        conn = get_connection_for_user(usuario)
        init_db(conn)
        case_selected = st.session_state.get("abono_case")
        caso_id_selected = case_selected[0] if isinstance(case_selected, tuple) else case_selected
        fecha_val = st.session_state.get("abono_fecha", date.today())
        if isinstance(fecha_val, str):
            try:
                fecha_val = pd.to_datetime(fecha_val).date()
            except Exception:
                fecha_val = date.today()
        add_abono(
            conn,
            fecha_val,
            st.session_state.get("abono_monto", 0.0),
            caso_id_selected,
            st.session_state.get("abono_obs", ""),
            creado_por=usuario,
        )
        # clear fields
        st.session_state["abono_monto"] = 0.0
        st.session_state["abono_obs"] = ""
        st.session_state["abono_fecha"] = date.today()
        # reset select to first option in next run
        st.session_state["reset_abono_case"] = True
        st.session_state["feedback"] = "Abono agregado correctamente."
        try:
            st.experimental_rerun()
        except Exception:
            pass
    except Exception as e:
        logging.exception("Error agregando abono (callback)")
        st.session_state["feedback"] = f"Error agregando abono: {e}"


# ------------------ Main ------------------


def main():
    st.set_page_config(page_title="Control de Abonos - Dashboard", layout="wide")

    if "logged_in" not in st.session_state:
        st.session_state["logged_in"] = False
        st.session_state["usuario"] = None

    # LOGIN
    if not st.session_state["logged_in"]:
        st.title("🔐 Acceso")
        if "credentials" not in st.secrets:
            st.error("Aplicación no configurada: falta la sección [credentials] en los secretos.")
            st.stop()
        user = st.text_input("Usuario", key="login_user")
        password = st.text_input("Contraseña", type="password", key="login_password")
        if st.button("Iniciar sesión", key="btn_login"):
            if check_password(user, password):
                st.session_state["logged_in"] = True
                st.session_state["usuario"] = user
                st.success(f"Bienvenido, {user} ✅")
            else:
                st.error("Usuario o contraseña incorrectos.")
        if not st.session_state["logged_in"]:
            st.stop()

    usuario = st.session_state.get("usuario")
    conn = get_connection_for_user(usuario)
    init_db(conn)

    st.markdown("""
    <style>
        .big-title { font-size:28px; font-weight:700; color:#0b3d91; }
        .subtle { color: #4b5563; }
    </style>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns([1, 4])
    with col1:
        st.button("Cerrar sesión", on_click=lambda: logout(), key="btn_logout")
    with col2:
        st.markdown(f'<div class="big-title">⚖️ Control de Abonos — Dashboard</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="subtle">Conectado como: {usuario}</div>', unsafe_allow_html=True)
    st.write("---")

    # init session keys (do NOT overwrite after widget creation)
    st.session_state.setdefault("new_cliente", "")
    st.session_state.setdefault("new_valor", 0.0)
    st.session_state.setdefault("new_descripcion", "")
    st.session_state.setdefault("new_etapa", "")
    st.session_state.setdefault("new_obs", "")
    st.session_state.setdefault("abono_case", None)
    st.session_state.setdefault("abono_fecha", date.today())
    st.session_state.setdefault("abono_monto", 0.0)
    st.session_state.setdefault("abono_obs", "")
    st.session_state.setdefault("feedback", "")

    # fetch fresh
    casos_df = fetch_casos(conn)
    abonos_df = fetch_abonos(conn)

    tab_casos, tab_abonos, tab_resumen, tab_reportes = st.tabs(["Casos", "Abonos", "Resumen", "Reportes"])

    # Feedback banner
    if st.session_state.get("feedback"):
        st.info(st.session_state.get("feedback"))
        # clear feedback after showing once
        st.session_state["feedback"] = ""

    # ---------- CASOS ----------
    with tab_casos:
        st.subheader("📁 Casos")
        st.markdown("Agregar nuevo caso (pulsa el botón 'Agregar Caso' para enviar).")

        col_a, col_b = st.columns(2)
        with col_a:
            st.text_input("Cliente", key="new_cliente", value=st.session_state["new_cliente"])
            st.number_input("Valor acordado", min_value=0.0, step=100.0, format="%.2f", key="new_valor", value=st.session_state["new_valor"])
        with col_b:
            st.text_input("Descripción", key="new_descripcion", value=st.session_state["new_descripcion"])
            st.text_input("Etapa", key="new_etapa", value=st.session_state["new_etapa"])
        st.text_area("Observaciones", key="new_obs", value=st.session_state["new_obs"])

        st.button("Agregar Caso", key="btn_add_caso", on_click=submit_new_case, args=(usuario,))

        # lista y edición
        casos_now = fetch_casos(conn)
        if not casos_now.empty:
            st.markdown("### Lista de casos")
            st.dataframe(casos_now, width="stretch")
            st.markdown("#### Editar / Eliminar caso")
            opciones_casos = [(int(r["id"]), f"{r['id']} — {r['cliente']} — {r['descripcion'] or ''}") for _, r in casos_now.iterrows()]
            seleccionado = st.selectbox("Selecciona caso", options=opciones_casos, format_func=lambda x: x[1], key="select_case_edit")
            caso_id_sel = seleccionado[0] if isinstance(seleccionado, tuple) else seleccionado

            with st.form("form_caso_edit"):
                c_row = casos_now.loc[casos_now["id"] == caso_id_sel].iloc[0]
                cliente_e = st.text_input("Cliente", value=c_row["cliente"], key="cliente_edit")
                descripcion_e = st.text_input("Descripción", value=c_row["descripcion"], key="desc_edit")
                valor_e = st.number_input("Valor acordado", value=float(c_row["valor_acordado"]), min_value=0.0, step=100.0, format="%.2f", key="valor_edit")
                etapa_e = st.text_input("Etapa", value=c_row["etapa"], key="etapa_edit")
                obs_e = st.text_area("Observaciones", value=c_row["observaciones"], key="obs_edit")
                btns = st.columns([1, 1])
                if btns[0].form_submit_button("Guardar cambios"):
                    try:
                        edit_caso(conn, caso_id_sel, cliente_e, descripcion_e, valor_e, etapa_e, obs_e)
                        st.success("Caso actualizado.")
                    except Exception:
                        logging.exception("Error editando caso")
                        st.error("Error al actualizar el caso. Revisa los logs.")
                confirm_delete = st.checkbox("Confirmo eliminación de este caso (y sus abonos).", key=f"confirm_case_{caso_id_sel}")
                if btns[1].form_submit_button("Eliminar caso"):
                    if not confirm_delete:
                        st.error("Marca la casilla de confirmación para eliminar.")
                    else:
                        try:
                            delete_caso(conn, caso_id_sel)
                            st.success("Caso eliminado.")
                        except Exception:
                            logging.exception("Error eliminando caso")
                            st.error("Error al eliminar el caso. Revisa los logs.")

    # ---------- ABONOS ----------
    with tab_abonos:
        st.subheader("💰 Abonos")
        casos_now = fetch_casos(conn)
        if casos_now.empty:
            st.info("Registra primero al menos un caso para agregar abonos.")
        else:
            st.markdown("Agregar nuevo abono (pulsa el botón 'Agregar Abono' para enviar).")
            opciones = [(int(r["id"]), f"{r['id']} — {r['cliente']} — {r['descripcion'] or ''}") for _, r in casos_now.iterrows()]

            # ensure default select option exists
            if opciones and st.session_state.get("abono_case") is None:
                st.session_state["abono_case"] = opciones[0]

            # compute default index safely
            default_index = 0
            try:
                stored = st.session_state.get("abono_case")
                stored_id = stored[0] if isinstance(stored, tuple) else stored
                default_index = next((i for i, o in enumerate(opciones) if o[0] == stored_id), 0)
            except Exception:
                default_index = 0

            st.selectbox("Selecciona Caso", options=opciones, format_func=lambda x: x[1], index=default_index, key="abono_case")
            st.date_input("Fecha", value=st.session_state["abono_fecha"], key="abono_fecha")
            st.number_input("Monto", min_value=0.0, step=100.0, format="%.2f", key="abono_monto", value=st.session_state["abono_monto"])
            st.text_area("Observaciones", key="abono_obs", value=st.session_state["abono_obs"])

            st.button("Agregar Abono", key="btn_add_abono", on_click=submit_new_abono, args=(usuario,))

        # Mostrar abonos y edición
        abonos = fetch_abonos(conn)
        if not abonos.empty:
            st.markdown("### Últimos abonos")
            st.dataframe(abonos, width="stretch")

            st.markdown("#### Editar / Eliminar abono")
            opciones_abonos = [(int(r["id"]), f"{r['id']} — {r['cliente']} — {r['fecha']} — ${float(r['monto']):,.2f}") for _, r in abonos.iterrows()]
            elegido = st.selectbox("Selecciona abono", options=opciones_abonos, format_func=lambda x: x[1], key="select_abono_edit")
            abono_id_sel = elegido[0] if isinstance(elegido, tuple) else elegido

            with st.form("form_abono_edit"):
                a_row = abonos.loc[abonos["id"] == abono_id_sel].iloc[0]
                caso_index = [o[0] for o in opciones].index(int(a_row["caso_id"])) if opciones else 0
                st.selectbox("Caso (editar)", options=opciones, format_func=lambda x: x[1], index=caso_index, key="case_edit_abono")
                st.date_input("Fecha", value=pd.to_datetime(a_row["fecha"]).date(), key="fecha_edit")
                st.number_input("Monto", value=float(a_row["monto"]), min_value=0.0, step=100.0, format="%.2f", key="monto_edit")
                st.text_area("Observaciones", value=a_row["observaciones"], key="obs_abono_edit")
                btns_ab = st.columns([1, 1])
                if btns_ab[0].form_submit_button("Guardar cambios"):
                    try:
                        edit_abono(conn, abono_id_sel, st.session_state["fecha_edit"], st.session_state["monto_edit"], st.session_state["case_edit_abono"][0], st.session_state["obs_abono_edit"])
                        st.success("Abono actualizado.")
                    except Exception:
                        logging.exception("Error editando abono")
                        st.error("Error al actualizar el abono. Revisa los logs.")
                confirm_delete_ab = st.checkbox("Confirmo eliminación de este abono.", key=f"confirm_ab_{abono_id_sel}")
                if btns_ab[1].form_submit_button("Eliminar abono"):
                    if not confirm_delete_ab:
                        st.error("Marca la casilla de confirmación para eliminar.")
                    else:
                        try:
                            delete_abono(conn, abono_id_sel)
                            st.success("Abono eliminado.")
                        except Exception:
                            logging.exception("Error eliminando abono")
                            st.error("Error al eliminar el abono. Revisa los logs.")

    # ---------- RESUMEN ----------
    with tab_resumen:
        st.subheader("📊 Resumen por Caso")
        casos_all = fetch_casos(conn)
        clientes = ["Todos"] + sorted(list(casos_all["cliente"].dropna().unique())) if not casos_all.empty else ["Todos"]
        etapas = ["Todos"] + sorted(list(casos_all["etapa"].fillna("").unique()))
        cliente_filter = st.selectbox("Filtrar por cliente", clientes, key="filter_cliente")
        etapa_filter = st.selectbox("Filtrar por etapa", etapas, key="filter_etapa")

        resumen_df = resumen_por_caso(conn, cliente_filter=cliente_filter, etapa_filter=etapa_filter)
        if resumen_df.empty:
            st.info("No hay datos disponibles con los filtros seleccionados.")
        else:
            total_acordado = resumen_df["valor_acordado"].sum()
            total_abonado = resumen_df["total_abonado"].sum()
            total_pendiente = resumen_df["saldo_pendiente"].sum()
            colA, colB, colC = st.columns(3)
            colA.metric("Total valor acordado", money(total_acordado))
            colB.metric("Total abonado", money(total_abonado))
            colC.metric("Total saldo pendiente", money(total_pendiente))
            display = resumen_df.copy()
            display["estado"] = display["saldo_pendiente"].apply(lambda x: "Pendiente" if x > 0.0 else "Pagado")
            display["valor_acordado"] = display["valor_acordado"].apply(money)
            display["total_abonado"] = display["total_abonado"].apply(money)
            display["saldo_pendiente"] = display["saldo_pendiente"].apply(money)
            st.dataframe(display, width="stretch")
            try:
                chart_df = resumen_df.set_index("descripcion")[["saldo_pendiente"]].sort_values("saldo_pendiente", ascending=False)
                st.bar_chart(chart_df, height=300)
            except Exception:
                chart_df = resumen_df.set_index("cliente")[["saldo_pendiente"]].sort_values("saldo_pendiente", ascending=False)
                st.bar_chart(chart_df, height=300)
            st.download_button("⬇️ Exportar Resumen a CSV", data=to_csv_bytes(resumen_df), file_name="resumen_abonos.csv", mime="text/csv")
            st.download_button("⬇️ Exportar Resumen a Excel", data=to_excel_bytes(resumen_df), file_name="resumen_abonos.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ---------- REPORTES ----------
    with tab_reportes:
        st.subheader("📑 Reportes y Exportes globales")
        df_export = resumen_por_caso(conn)
        if df_export.empty:
            st.info("No hay datos para exportar.")
        else:
            total_acordado = df_export["valor_acordado"].sum()
            total_abonado = df_export["total_abonado"].sum()
            total_pendiente = df_export["saldo_pendiente"].sum()
            r1, r2, r3 = st.columns(3)
            r1.metric("Total valor acordado", money(total_acordado))
            r2.metric("Total abonado", money(total_abonado))
            r3.metric("Total saldo pendiente", money(total_pendiente))
            st.dataframe(df_export.assign(
                valor_acordado=df_export["valor_acordado"].apply(money),
                total_abonado=df_export["total_abonado"].apply(money),
                saldo_pendiente=df_export["saldo_pendiente"].apply(money),
            ), width="stretch")
            st.download_button("⬇️ Exportar CSV (Global)", data=to_csv_bytes(df_export), file_name="resumen_abonos_global.csv", mime="text/csv")
            st.download_button("⬇️ Exportar Excel (Global)", data=to_excel_bytes(df_export), file_name="resumen_abonos_global.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def logout():
    st.session_state["logged_in"] = False
    st.session_state["usuario"] = None


if __name__ == "__main__":
    main()
